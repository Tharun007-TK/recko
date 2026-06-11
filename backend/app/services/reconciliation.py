"""
Reconciliation service layer
Handles reading, mapping, normalizing, comparing, and reporting on records
"""

from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
import pandas as pd
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


@dataclass
class ReconciliationResult:
    """Result of reconciliation operation"""

    job_id: str
    total_records: int
    matched: int
    mismatched: int
    missing_in_gst: int
    missing_in_tally: int
    format_differences: int
    mismatches: List[Dict[str, Any]]


class ReconciliationService:
    """Service for reconciling Excel files"""

    def __init__(self, mapping_profile: Dict[str, Any], rule_profile: Dict[str, Any]):
        """Initialize with mapping and rule profiles"""
        self.mapping_profile = mapping_profile or {}
        self.rule_profile = rule_profile or {}
        self.mappings = self.mapping_profile.get("mapping_json", [])
        self.rules = self.rule_profile.get("rules_json", {})

    # ─── File Loading ─────────────────────────────────────────────────────

    def _detect_header_row(self, df_raw: "pd.DataFrame", max_scan: int = 15) -> int:
        """Auto-detect the header row by finding the first row with most non-null string cells."""
        best_row = 0
        best_score = 0
        for i in range(min(max_scan, len(df_raw))):
            row = df_raw.iloc[i]
            score = sum(1 for v in row if isinstance(v, str) and v.strip() and not v.lower().startswith('unnamed'))
            if score > best_score:
                best_score = score
                best_row = i
        return best_row

    def load_excel_file(self, file_data: bytes, sheet_name: int = 0) -> Optional[pd.DataFrame]:
        """Load Tally Excel file into DataFrame with auto-detected header row."""
        try:
            raw = pd.read_excel(file_data, sheet_name=sheet_name, header=None)
            header_row = self._detect_header_row(raw)
            df = pd.read_excel(file_data, sheet_name=sheet_name, header=header_row)
            # Drop completely empty rows and columns
            df = df.dropna(how='all').dropna(axis=1, how='all')
            logger.info(f"Loaded Excel (sheet={sheet_name}, header_row={header_row}): {len(df)} rows, {len(df.columns)} cols")
            return df
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return None

    def load_gst_file(self, file_data: bytes) -> Optional[pd.DataFrame]:
        """Load GSTR-2B Excel file, selecting the B2B sheet and auto-detecting the header row."""
        try:
            all_sheets = pd.read_excel(file_data, sheet_name=None, header=None)
            # Priority: B2B sheet, otherwise the largest sheet
            target = None
            for preferred in ['B2B', 'b2b', 'Sheet1', 'Sheet']:
                if preferred in all_sheets:
                    target = preferred
                    break
            if target is None:
                # Fall back to the sheet with the most rows
                target = max(all_sheets.keys(), key=lambda s: len(all_sheets[s]))
            raw = all_sheets[target]
            header_row = self._detect_header_row(raw)
            df = pd.read_excel(file_data, sheet_name=target, header=header_row)
            df = df.dropna(how='all').dropna(axis=1, how='all')
            logger.info(f"Loaded GST file (sheet='{target}', header_row={header_row}): {len(df)} rows, {len(df.columns)} cols")
            return df
        except Exception as e:
            logger.error(f"Error loading GST file: {e}")
            return None


    # ─── Data Normalization ───────────────────────────────────────────────

    def normalize_value(self, value: Any) -> str:
        """Normalize a value according to rule profile"""
        if pd.isna(value):
            return ""

        # Convert to string
        if isinstance(value, float) and value.is_integer():
            normalized = str(int(value)).strip()
        else:
            normalized = str(value).strip()

        # Trim spaces
        if self.rules.get("trim_spaces", True):
            normalized = normalized.strip()

        # Ignore case
        if self.rules.get("ignore_case", False):
            normalized = normalized.lower()

        # Normalize dates
        if self.rules.get("normalize_dates", False):
            normalized = self._normalize_date(normalized)

        # Remove separators
        if self.rules.get("remove_separators", False):
            normalized = self._remove_separators(normalized)

        # Numeric rounding
        if self.rules.get("numeric_rounding") is not None:
            normalized = self._apply_numeric_rounding(
                normalized, self.rules["numeric_rounding"]
            )

        return normalized

    def _normalize_date(self, value: str) -> str:
        """Normalize date to YYYY-MM-DD format"""
        try:
            # Try parsing common date formats
            date_formats = [
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d.%m.%Y",
            ]
            for fmt in date_formats:
                try:
                    parsed = datetime.strptime(value, fmt)
                    return parsed.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return value
        except Exception as e:
            logger.warning(f"Error normalizing date: {e}")
            return value

    def _remove_separators(self, value: str) -> str:
        """Remove common separators"""
        separators = [",", "-", "/", ".", " "]
        for sep in separators:
            value = value.replace(sep, "")
        return value

    def _apply_numeric_rounding(self, value: str, decimal_places: int) -> str:
        """Apply numeric rounding"""
        try:
            if not value:
                return value
            # Remove non-numeric characters except decimal point
            numeric_str = "".join(c for c in value if c.isdigit() or c == ".")
            if not numeric_str:
                return value
            numeric_val = float(numeric_str)
            rounded = round(numeric_val, decimal_places)
            return str(rounded)
        except Exception as e:
            logger.warning(f"Error applying numeric rounding: {e}")
            return value

    # ─── Standardize and Clean Data ────────────────────────────────────────

    def _standardize_and_clean(self, df: pd.DataFrame, is_gst: bool = False) -> pd.DataFrame:
        """Standardize column names and clean data for reconciliation."""
        df = df.copy()
        
        # 1. Standardize column names based on common headers
        column_map = {}
        for col in df.columns:
            col_str = str(col).strip().lower()
            
            # invoice_no
            if col_str in ['invoice number', 'invoice no', 'invoice_number', 'invoice_no', 'invoice', 'voucher no', 'voucher number']:
                column_map[col] = 'invoice_no'
            # invoice_date
            elif col_str in ['invoice date', 'date', 'invoice_date', 'voucher date']:
                column_map[col] = 'invoice_date'
            # tax_amount
            elif col_str in ['tax', 'tax amount', 'tax_amount', 'total tax', 'integrated tax', 'central tax', 'state/ut tax']:
                # If multiple tax columns exist, this might map the last one or we can map them specifically.
                # Since the prompt specifies mapping "Tax" / "Tax Amount" -> tax_amount, we do that.
                if 'tax_amount' not in column_map.values(): # Map only the first one found or we could sum them
                    column_map[col] = 'tax_amount'
            # taxable_value
            elif col_str in ['taxable value', 'taxable_value', 'taxable']:
                column_map[col] = 'taxable_value'
            # gstin
            elif col_str in ['supplier gstin', 'gstin', 'gstin/uin', 'gstin/uin of supplier', 'ctin']:
                column_map[col] = 'gstin'
        
        # Also let's check for specific tax columns if a general one is not found
        df = df.rename(columns=column_map)
        
        # Sum specific tax columns if a single tax amount isn't defined
        if 'tax_amount' not in df.columns:
            tax_cols = [c for c in df.columns if str(c).strip().lower() in ['integrated tax', 'central tax', 'state/ut tax', 'integrated tax amount', 'central tax amount', 'state/ut tax amount']]
            if tax_cols:
                for c in tax_cols:
                    df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
                df['tax_amount'] = df[tax_cols].sum(axis=1)

        # Strip all strings
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
        
        # Ensure target columns exist
        for target in ['invoice_no', 'invoice_date', 'tax_amount', 'taxable_value', 'gstin']:
            if target not in df.columns:
                df[target] = None
        
        # Convert numeric columns safely
        df['taxable_value'] = pd.to_numeric(df['taxable_value'], errors='coerce').fillna(0.0)
        df['tax_amount'] = pd.to_numeric(df['tax_amount'], errors='coerce').fillna(0.0)
        
        # Format date safely
        def safe_date(d):
            if pd.isna(d) or d == 'nan' or not str(d).strip():
                return ""
            try:
                parsed = pd.to_datetime(d)
                return parsed.strftime("%Y-%m-%d")
            except Exception:
                return str(d).strip()
                
        df['invoice_date'] = df['invoice_date'].apply(safe_date)
        
        # Ensure invoice_no is strictly cleaned and stripped to prevent trailing space mismatches
        df['invoice_no'] = df['invoice_no'].astype(str).str.strip()
        
        # Best practice: Also strip the GSTIN column just in case
        if 'gstin' in df.columns:
            df['gstin'] = df['gstin'].astype(str).str.strip().str.upper()
            
        return df

    # ─── Reconciliation Logic ──────────────────────────────────────────────

    def reconcile(
        self, tally_df: pd.DataFrame, gst_df: pd.DataFrame, job_id: str, firm_id: str
    ) -> ReconciliationResult:
        """Reconcile two DataFrames"""
        
        tally_clean = self._standardize_and_clean(tally_df, is_gst=False)
        gst_clean = self._standardize_and_clean(gst_df, is_gst=True)
        
        # Remove empty invoice_no
        tally_clean = tally_clean[~tally_clean['invoice_no'].isin(['nan', 'None', ''])]
        gst_clean = gst_clean[~gst_clean['invoice_no'].isin(['nan', 'None', ''])]

        # Aggregate by invoice_no to handle multiple line items per invoice
        tally_agg = tally_clean.groupby('invoice_no', as_index=False).agg({
            'taxable_value': 'sum',
            'tax_amount': 'sum',
            'invoice_date': 'first',
            'gstin': 'first'
        })
        gst_agg = gst_clean.groupby('invoice_no', as_index=False).agg({
            'taxable_value': 'sum',
            'tax_amount': 'sum',
            'invoice_date': 'first',
            'gstin': 'first'
        })

        # Determine total unique invoice numbers across both sets
        all_invoices = set(tally_agg['invoice_no']).union(set(gst_agg['invoice_no']))
        total_records = len(all_invoices)

        # Merge DataFrames (Left=GST, Right=Tally as requested)
        merged_df = pd.merge(gst_agg, tally_agg, on='invoice_no', how='outer', suffixes=('_gst', '_tally'), indicator=True)
        
        matched = 0
        mismatched = 0
        missing_in_tally = 0
        missing_in_gst = 0
        format_differences = 0
        
        mismatches: List[Dict[str, Any]] = []

        import math

        for _, row in merged_df.iterrows():
            merge_status = row['_merge']
            invoice_no = row['invoice_no']
            
            if merge_status == 'both':
                taxable_gst = row['taxable_value_gst']
                taxable_tally = row['taxable_value_tally']
                tax_gst = row['tax_amount_gst']
                tax_tally = row['tax_amount_tally']
                
                # Check for mismatches
                is_mismatch = False
                field_mismatches = []
                
                if not math.isclose(taxable_gst, taxable_tally, abs_tol=0.01):
                    is_mismatch = True
                    field_mismatches.append({
                        "job_id": job_id,
                        "firm_id": firm_id,
                        "category": "field_mismatch",
                        "match_key": invoice_no,
                        "field_name": "taxable_value",
                        "tally_value": str(taxable_tally),
                        "gst_value": str(taxable_gst),
                        "normalized_tally": str(taxable_tally),
                        "normalized_gst": str(taxable_gst),
                        "reason": f"Taxable Value mismatch: {taxable_tally} vs {taxable_gst}"
                    })
                    
                if not math.isclose(tax_gst, tax_tally, abs_tol=0.01):
                    is_mismatch = True
                    field_mismatches.append({
                        "job_id": job_id,
                        "firm_id": firm_id,
                        "category": "field_mismatch",
                        "match_key": invoice_no,
                        "field_name": "tax_amount",
                        "tally_value": str(tax_tally),
                        "gst_value": str(tax_gst),
                        "normalized_tally": str(tax_tally),
                        "normalized_gst": str(tax_gst),
                        "reason": f"Tax Amount mismatch: {tax_tally} vs {tax_gst}"
                    })
                    
                if is_mismatch:
                    mismatched += 1
                    format_differences += len(field_mismatches)
                    mismatches.extend(field_mismatches)
                else:
                    matched += 1
                    
            elif merge_status == 'left_only':
                # Missing in Tally (Present in GST)
                missing_in_tally += 1
                taxable_gst = row['taxable_value_gst']
                tax_gst = row['tax_amount_gst']
                mismatches.append({
                    "job_id": job_id,
                    "firm_id": firm_id,
                    "category": "missing_in_tally",
                    "match_key": invoice_no,
                    "field_name": "",
                    "tally_value": "",
                    "gst_value": f"Taxable: {taxable_gst}, Tax: {tax_gst}",
                    "normalized_tally": "",
                    "normalized_gst": str(taxable_gst),
                    "reason": "Record exists in GST but not in Tally",
                })
            elif merge_status == 'right_only':
                # Missing in GST (Present in Tally)
                missing_in_gst += 1
                taxable_tally = row['taxable_value_tally']
                tax_tally = row['tax_amount_tally']
                mismatches.append({
                    "job_id": job_id,
                    "firm_id": firm_id,
                    "category": "missing_in_gst",
                    "match_key": invoice_no,
                    "field_name": "",
                    "tally_value": f"Taxable: {taxable_tally}, Tax: {tax_tally}",
                    "gst_value": "",
                    "normalized_tally": str(taxable_tally),
                    "normalized_gst": "",
                    "reason": "Record exists in Tally but not in GST",
                })

        return ReconciliationResult(
            job_id=job_id,
            total_records=total_records,
            matched=matched,
            mismatched=mismatched,
            missing_in_gst=missing_in_gst,
            missing_in_tally=missing_in_tally,
            format_differences=format_differences,
            mismatches=mismatches,
        )

    def generate_excel_report(self, result: ReconciliationResult) -> bytes:
        """Generate an Excel report for the reconciliation result"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        
        # Helper to style headers
        def style_header(ws):
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Count"])
        ws_summary.append(["Job ID", result.job_id])
        ws_summary.append(["Total Records Checked", result.total_records])
        ws_summary.append(["Matched", result.matched])
        ws_summary.append(["Mismatched", result.mismatched])
        ws_summary.append(["Missing in GST", result.missing_in_gst])
        ws_summary.append(["Missing in Tally", result.missing_in_tally])
        ws_summary.append(["Format Differences", result.format_differences])
        style_header(ws_summary)
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40

        # Create separate data lists for each category
        field_mismatches = []
        missing_in_gst = []
        missing_in_tally = []

        for m in result.mismatches:
            row = [
                m.get("match_key", ""),
                m.get("field_name", ""),
                m.get("tally_value", ""),
                m.get("gst_value", ""),
                m.get("normalized_tally", ""),
                m.get("normalized_gst", ""),
                m.get("reason", "")
            ]
            if m.get("category") == "field_mismatch":
                field_mismatches.append(row)
            elif m.get("category") == "missing_in_gst":
                missing_in_gst.append(row)
            elif m.get("category") == "missing_in_tally":
                missing_in_tally.append(row)

        headers = [
            "Match Key", "Field Name", "Raw Tally Value", "Raw GST Value", 
            "Normalized Tally", "Normalized GST", "Reason"
        ]

        # 2. Field Mismatches Sheet
        ws_mismatches = wb.create_sheet(title="Field Mismatches")
        ws_mismatches.append(headers)
        for row in field_mismatches:
            ws_mismatches.append(row)
        style_header(ws_mismatches)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_mismatches.column_dimensions[col].width = 20

        # 3. Missing in GST Sheet
        ws_missing_gst = wb.create_sheet(title="Missing in GST")
        ws_missing_gst.append(headers)
        for row in missing_in_gst:
            ws_missing_gst.append(row)
        style_header(ws_missing_gst)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_missing_gst.column_dimensions[col].width = 20

        # 4. Missing in Tally Sheet
        ws_missing_tally = wb.create_sheet(title="Missing in Tally")
        ws_missing_tally.append(headers)
        for row in missing_in_tally:
            ws_missing_tally.append(row)
        style_header(ws_missing_tally)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_missing_tally.column_dimensions[col].width = 20

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
